import openpyxl
from bs4 import BeautifulSoup
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime

# URL der Zielwebsite
url = "https://www.interhyp.de/immobilienpreise/#immobilienpreisentwicklung"

# Selenium WebDriver einrichten
driver = webdriver.Chrome()
wait = WebDriverWait(driver, 30)  # Erhöhte Wartezeit für langsame Ladezeiten

# Excel-Datei einrichten
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Immobilienpreise"
sheet.append(["PLZ", "Zeitraum", "Marktwert Haus", "Marktwert Wohnung", "Marktwert Haus Aktuell", "Marktwert Wohnung Aktuell"])

cookie_banner_closed = False

# Funktion zum Loggen mit Zeitstempel
def log_with_timestamp(message):
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {message}")

def select_dropdown_option(option_text):
    for attempt in range(3):  # Mehrere Versuche, falls ein Fehler auftritt
        try:
            dropdown = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "input[aria-haspopup='listbox']")))
            driver.execute_script("arguments[0].scrollIntoView(true);", dropdown)  # Scrollen, um das Dropdown sichtbar zu machen
            dropdown.click()
            time.sleep(2)  # Zeit für das Dropdown-Menü zum Öffnen
            option = wait.until(EC.element_to_be_clickable((By.XPATH, f"//span[text()='{option_text}']")))
            option.click()
            log_with_timestamp(f"Dropdown-Option '{option_text}' ausgewählt.")
            return
        except Exception as e:
            log_with_timestamp(f"Fehler beim Auswählen der Dropdown-Option '{option_text}': {e}")
            time.sleep(5)  # Wartezeit vor erneutem Versuch

    log_with_timestamp(f"Dropdown-Option '{option_text}' konnte nach mehreren Versuchen nicht ausgewählt werden.")

def scrape_data_selenium(plz):
    """Daten für eine einzelne PLZ mit Selenium abfragen."""
    global cookie_banner_closed

    driver.get(url)
    time.sleep(10)  # Warten, bis die Seite vollständig geladen ist

    if not cookie_banner_closed:
        try:
            cookie_button = wait.until(EC.element_to_be_clickable((By.ID, "cm-btnAcceptAll")))
            cookie_button.click()
            log_with_timestamp("Cookie-Banner geschlossen.")
            cookie_banner_closed = True
        except Exception as e:
            log_with_timestamp("Kein Cookie-Banner gefunden oder Fehler beim Schließen: " + str(e))

    try:
        # Eingabe der PLZ
        input_field = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "input[placeholder='Postleitzahl / Ort']")))
        input_field.clear()
        input_field.send_keys(Keys.CONTROL + "a")  # Sicherstellen, dass alles gelöscht wird
        input_field.send_keys(Keys.DELETE)  # Löschen der Inhalte
        input_field.send_keys(plz)
        input_field.send_keys(Keys.RETURN)

        # Wartezeit für die Aktualisierung der Daten
        time.sleep(12)

        # Prüfen, ob genügend Daten vorliegen
        try:
            error_message = driver.find_element(By.XPATH, "//h5[contains(text(), 'Ergebnisanzeige nicht möglich')]")
            if error_message:
                log_with_timestamp(f"Keine Daten für PLZ {plz}. Überspringen...")
                sheet.append([plz, "N/A", "N/A", "N/A", "N/A", "N/A"])
                workbook.save("Immobilienpreise2.xlsx")
                return
        except:
            pass

        # Extrahiere die aktuellen Werte
        try:
            haus_div = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, "content_LXx3p4L4YM")))[0]
            haus_soup = BeautifulSoup(haus_div.get_attribute('innerHTML'), 'html.parser')
            marktwert_haus_aktuell = haus_soup.select_one("p:-soup-contains('Aktuell:') strong").get_text(strip=True)
        except Exception as e:
            marktwert_haus_aktuell = "N/A"
            log_with_timestamp(f"Fehler beim Auslesen des aktuellen Marktwerts Haus für PLZ {plz}: {e}")

        try:
            wohnung_div = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, "content_LXx3p4L4YM")))[1]
            wohnung_soup = BeautifulSoup(wohnung_div.get_attribute('innerHTML'), 'html.parser')
            marktwert_wohnung_aktuell = wohnung_soup.select_one("p:-soup-contains('Aktuell:') strong").get_text(strip=True)
        except Exception as e:
            marktwert_wohnung_aktuell = "N/A"
            log_with_timestamp(f"Fehler beim Auslesen des aktuellen Marktwerts Wohnung für PLZ {plz}: {e}")

        # Werte für verschiedene Zeiträume aus Dropdown
        zeitraeume = ["1 Jahr", "2 Jahre", "5 Jahre", "10 Jahre"]
        for zeitraum in zeitraeume:
            select_dropdown_option(zeitraum)
            time.sleep(12)  # Erhöhte Zeit für die Aktualisierung der Werte

            try:
                haus_div = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, "content_LXx3p4L4YM")))[0]
                haus_soup = BeautifulSoup(haus_div.get_attribute('innerHTML'), 'html.parser')
                marktwert_haus = haus_soup.select_one("p strong").get_text(strip=True)  # Dynamisch das erste <strong> Tag holen
            except Exception as e:
                marktwert_haus = "N/A"
                log_with_timestamp(f"Fehler beim Auslesen der Marktwerte Haus für PLZ {plz} und Zeitraum {zeitraum}: {e}")

            try:
                wohnung_div = wait.until(EC.presence_of_all_elements_located((By.CLASS_NAME, "content_LXx3p4L4YM")))[1]
                wohnung_soup = BeautifulSoup(wohnung_div.get_attribute('innerHTML'), 'html.parser')
                marktwert_wohnung = wohnung_soup.select_one("p strong").get_text(strip=True)  # Dynamisch das erste <strong> Tag holen
            except Exception as e:
                marktwert_wohnung = "N/A"
                log_with_timestamp(f"Fehler beim Auslesen der Marktwerte Wohnung für PLZ {plz} und Zeitraum {zeitraum}: {e}")

            # Daten in Excel schreiben und speichern
            sheet.append([plz, zeitraum, marktwert_haus, marktwert_wohnung, marktwert_haus_aktuell, marktwert_wohnung_aktuell])
            workbook.save("Immobilienpreise2.xlsx")

    except Exception as e:
        sheet.append([plz, "N/A", "N/A", "N/A", "N/A", "N/A"])
        workbook.save("Immobilienpreise2.xlsx")
        log_with_timestamp(f"Fehler bei PLZ {plz}: {e}")

# PLZ aus Excel-Datei einlesen
def read_plz_from_excel(file_path, start_plz):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    plz_list = [str(row[0].value).zfill(5) for row in sheet.iter_rows(min_row=2, max_col=1)]
    if start_plz in plz_list:
        start_index = plz_list.index(start_plz) + 1
        return plz_list[start_index:]
    return []

# Pfad zur Excel-Datei mit PLZ
input_excel_file = "PLZ_Liste.xlsx"

# PLZ-Liste aus Excel einlesen
deutsche_plz = read_plz_from_excel(input_excel_file)

# Alle PLZ durchlaufen
for plz in deutsche_plz:
    try:
        log_with_timestamp(f"Abfrage für PLZ {plz}...")
        scrape_data_selenium(plz)
    except Exception as e:
        log_with_timestamp(f"Fehler bei PLZ {plz}: {e}")

# Selenium WebDriver schließen
driver.quit()

log_with_timestamp(f"Daten wurden erfolgreich in 'Immobilienpreise2.xlsx' gespeichert.")
